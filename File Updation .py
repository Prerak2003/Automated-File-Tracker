#!/usr/bin/env python
# coding: utf-8

# In[ ]:


import os
from openpyxl import Workbook, load_workbook
from datetime import datetime
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.base import MIMEBase
from email import encoders

# Define the constant directory path
DIRECTORY_PATH = r"C:\Users\prera\Desktop\ANANT\GSTR 1 & 3B"

def list_files(directory):
    files = []
    for root, _, filenames in os.walk(directory):
        for filename in filenames:
            files.append(os.path.join(root, filename))
    return files

def update_excel(current_data_filename):
    directory = DIRECTORY_PATH  # Use the predefined directory path
    print("Updating Excel in directory:", directory)

    current_files = list_files(directory)
    print("Current files:", current_files)

    current_data_sheet = 'Current Data'
    report_data_sheet = 'Report Data'

    current_data_filepath = os.path.join(directory, current_data_filename)

    if os.path.exists(current_data_filepath):
        print("Opening existing workbook...")
        wb = load_workbook(current_data_filepath)
        current_data = wb[current_data_sheet]
        report_data = wb[report_data_sheet]
    else:
        print("Creating new workbook...")
        wb = Workbook()
        current_data = wb.active
        current_data.title = current_data_sheet
        report_data = wb.create_sheet(title=report_data_sheet)
        report_data.append(['New Files:'])

    old_files = [cell.value for cell in current_data['A'] if cell.value]
    print("Old files:", old_files)

    new_files = [file for file in current_files if file not in old_files and file != current_data_filepath]
    print("New files:", new_files)

    # Clear current data sheet before updating
    current_data.delete_rows(1, current_data.max_row)

    # Update report data sheet with new files
    if new_files:
        for file in new_files:
            report_data.append([file])

    # Update current data sheet with new files only
    for file in new_files:
        current_data.append([file])

    wb.save(current_data_filepath)
    print("Excel update complete.")

    # Email the updated Excel file
    send_email(current_data_filepath)

def send_email(filepath):
    sender_email = "mmnautomationdelhi@gmail.com"
    receiver_email = "anant.jain@mmnissim.com"
    password = "mmuwaddxmhrpcbkn"  # Your app password here

    msg = MIMEMultipart()
    msg['From'] = sender_email
    msg['To'] = receiver_email
    msg['Subject'] = "Updated Excel File"

    part = MIMEBase('application', "octet-stream")
    part.set_payload(open(filepath, "rb").read())
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', 'attachment; filename="%s"' % os.path.basename(filepath))
    msg.attach(part)

    with smtplib.SMTP_SSL("smtp.gmail.com", 465) as server:
        server.login(sender_email, password)
        server.sendmail(sender_email, receiver_email, msg.as_string())

    print("Email sent successfully.")

def main():
    current_data_filename = "Current_Data.xlsx"
    update_excel(current_data_filename)

if __name__ == "__main__":
    main()


# In[ ]:




